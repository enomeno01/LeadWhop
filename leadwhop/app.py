"""LeadWhop — Streamlit UI.

Deploy: push to GitHub → connect on share.streamlit.io → add secrets.
Local:  streamlit run app.py
"""
import hmac
import io
import os
import tempfile

import pandas as pd
import streamlit as st

from leadwhop.pipeline import Pipeline, STAGES

# ── Streamlit Cloud Secrets → env vars ──────────────────────────────────────
for _key in ("OPENAI_API_KEY", "SERPER_API_KEY", "LUSHA_API_KEY"):
    if not os.environ.get(_key):
        try:
            val = st.secrets.get(_key)
            if val:
                os.environ[_key] = val
        except Exception:
            pass

# ── Password gate ────────────────────────────────────────────────────────────
# The password lives ONLY in Streamlit Secrets (or a local .env). There is no
# hard-coded fallback: in a public repo a default password is the same as no
# password at all, and anyone who read it could burn the owner's API credits.
# If APP_PASSWORD is not configured the app fails closed rather than open.
_APP_PASSWORD = os.environ.get("APP_PASSWORD")
if not _APP_PASSWORD:
    try:
        _APP_PASSWORD = st.secrets.get("APP_PASSWORD")
    except Exception:
        _APP_PASSWORD = None

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "login_attempts" not in st.session_state:
    st.session_state.login_attempts = 0

if not st.session_state.authenticated:
    st.markdown("# 🎯 LeadWhop")

    if not _APP_PASSWORD:
        st.error("**APP_PASSWORD is not configured — the app is locked.**")
        st.markdown(
            "Set it before using the app:\n\n"
            "- **Streamlit Cloud:** Settings → Secrets → add "
            "`APP_PASSWORD = \"your-password\"`\n"
            "- **Local:** add `APP_PASSWORD=your-password` to your `.env`"
        )
        st.stop()

    # Slow down brute-force attempts without locking the owner out.
    if st.session_state.login_attempts >= 5:
        st.error("Too many failed attempts. Reload the page to try again.")
        st.stop()

    st.markdown("Enter the password to continue.")
    pwd = st.text_input("Password", type="password", key="pwd_input")
    if st.button("Enter", type="primary"):
        # Constant-time comparison: a plain == leaks length/prefix via timing.
        if hmac.compare_digest(str(pwd), str(_APP_PASSWORD)):
            st.session_state.authenticated = True
            st.session_state.login_attempts = 0
            st.rerun()
        else:
            st.session_state.login_attempts += 1
            left = 5 - st.session_state.login_attempts
            st.error(f"Wrong password. {left} attempt(s) left.")
    st.stop()

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="LeadWhop",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] { background-color: #0a2f38; }
.block-container { padding-top: 2rem; }
.metric-card {
    background: #124E5B;
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    text-align: center;
}
.metric-card .value { font-size: 2.2rem; font-weight: 800; color: #F2B33D; }
.metric-card .label { font-size: 0.85rem; color: #9CC3CB; margin-top: 0.2rem; }
.stage-badge {
    display: inline-block;
    background: #1E6674;
    color: #F2B33D;
    border-radius: 20px;
    padding: 0.2rem 0.8rem;
    font-size: 0.8rem;
    font-weight: 700;
    margin-bottom: 0.3rem;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    import pathlib as _pl
    _logo = _pl.Path(__file__).parent / "docs" / "logo.svg"
    if _logo.exists():
        st.image(str(_logo), width="stretch")
    st.divider()

    # API keys — only shown when not already injected from Secrets
    missing = [k for k in ("OPENAI_API_KEY", "SERPER_API_KEY", "LUSHA_API_KEY")
               if not os.environ.get(k)]
    if missing:
        st.markdown("### 🔑 API Keys")
        st.caption("Keys stay in this session only — never written to disk.")
        label_map = {"OPENAI_API_KEY": "OpenAI",
                     "SERPER_API_KEY": "Serper.dev",
                     "LUSHA_API_KEY":  "Lusha"}
        for key in missing:
            val = st.text_input(label_map[key], type="password", key=key)
            if val:
                os.environ[key] = val
        st.divider()

    st.markdown("### ⚙️ Pipeline stages")
    stage_meta = {
        "websites":  ("1", "Find websites",        True),
        "qualify":   ("2", "Check product fit",     True),
        "contacts":  ("3", "Find buyers",           False),
        "phones":    ("4", "Enrich phones",         False),
        "export":    ("5", "Export to CRM",         False),
        "mail":      ("6", "Create personalized e-mails", False),
    }
    selected = []
    for stage, (num, label, default) in stage_meta.items():
        if st.checkbox(f"**{num}** — {label}", value=default, key=f"stage_{stage}"):
            selected.append(stage)

    # Dependency notes — export can run standalone on a file that already
    # contains contact columns (Email, Title, Name) from a previous run.
    if "export" in selected and "contacts" not in selected:
        st.info("ℹ️ Stage 5 alone: your file must already contain contact "
                "columns (Email, Title, Name) from a previous 1-3 run.")
    if "contacts" in selected and "qualify" not in selected:
        st.info("ℹ️ Tip: Stage 2 filters companies before spending Lusha credits.")

    event_name = ""
    if "export" in selected:
        st.divider()
        event_name = st.text_input("📅 Event name", placeholder="e.g. PLMA 2026")

    custom_icp_prompt = ""
    custom_mail_prompt = ""
    if "qualify" in selected or "mail" in selected:
        st.divider()
        with st.expander("⚙️ Advanced settings", expanded=False):
            if "qualify" in selected:
                st.caption("**Product fit** — extra instructions for the AI "
                           "(leave blank to use defaults)")
                custom_icp_prompt = st.text_area(
                    "ICP instructions",
                    placeholder="e.g. Include craft distilleries even if small.",
                    height=80, label_visibility="collapsed",
                )
            if "mail" in selected:
                st.caption("**Email** — extra instructions for the AI "
                           "(leave blank to use defaults)")
                custom_mail_prompt = st.text_area(
                    "Mail instructions",
                    placeholder="e.g. Mention our new matte black finish. Keep tone formal.",
                    height=80, label_visibility="collapsed",
                )

# ── Main ─────────────────────────────────────────────────────────────────────
st.markdown("# 🎯 LeadWhop")
st.caption("Company list in → CRM-ready qualified leads out")
st.divider()

uploaded = st.file_uploader(
    "📂 Upload Excel (.xlsx)",
    type=["xlsx"],
)
st.caption(
    "**Required column:** Company &nbsp;·&nbsp; "
    "**Optional:** Country, Website &nbsp;·&nbsp; "
    "**For stages 5-6 only:** Email, Name, Title, AI_Note"
)

if uploaded:
    preview = pd.read_excel(uploaded)
    st.markdown(f"**{len(preview)} companies loaded** — preview:")
    st.dataframe(preview.head(8), width="stretch", height=220)

    # Validation
    missing_keys = []
    if not os.environ.get("OPENAI_API_KEY"):  missing_keys.append("OpenAI")
    if not os.environ.get("SERPER_API_KEY"):  missing_keys.append("Serper")
    if ("contacts" in selected or "phones" in selected) and not os.environ.get("LUSHA_API_KEY"):
        missing_keys.append("Lusha")

    if missing_keys:
        st.warning(f"⚠️ Add missing API keys in the sidebar: {', '.join(missing_keys)}")
    elif not selected:
        st.info("Select at least one stage in the sidebar.")
    else:
        st.divider()
        run_col, _ = st.columns([1, 3])
        with run_col:
            run = st.button("▶️ Run pipeline", type="primary", width="stretch")

        if run:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(uploaded.getvalue())
                tmp_path = tmp.name

            status   = st.empty()
            progress = st.progress(0.0)

            def cb(stage, i, n):
                _, label, _ = stage_meta[stage]
                progress.progress(i / n, text=f"**{label}** — {i} / {n}")

            from leadwhop import status as lw_status
            lw_status.clear()
            try:
                pipe   = Pipeline.from_config("config/settings.yaml")
                result = pipe.run(tmp_path, stages=selected,
                                  event_name=event_name, progress_cb=cb,
                                  custom_icp_prompt=custom_icp_prompt,
                                  custom_mail_prompt=custom_mail_prompt)
                progress.progress(1.0, text="✅ Done")
                status.success("Pipeline complete!")

                # Hide internal debug columns from the user-facing output
                _hidden = ["Website_Debug"]
                result = result.drop(columns=[c for c in _hidden
                                              if c in result.columns])
            except Exception as exc:
                st.error(f"Pipeline error: {exc}")
                for w in lw_status.get_warnings():
                    st.error(f"🚨 {w}")
                st.stop()

            # API warnings (credits, auth, rate limits) — never silent
            for w in lw_status.get_warnings():
                st.error(f"🚨 {w}")

            st.divider()

            # ── Summary metrics ──────────────────────────────────────────
            total    = len(result)
            reviewed = int(result["Needs_Review"].sum()) if "Needs_Review" in result.columns else 0
            verified = total - reviewed
            with_email = int((result["Email"].notna() & (result["Email"] != "")).sum()) \
                if "Email" in result.columns else 0

            m1, m2, m3, m4 = st.columns(4)
            m1.markdown(f'<div class="metric-card"><div class="value">{total}</div>'
                        f'<div class="label">Total rows</div></div>', unsafe_allow_html=True)
            m2.markdown(f'<div class="metric-card"><div class="value">{verified}</div>'
                        f'<div class="label">Auto-verified</div></div>', unsafe_allow_html=True)
            m3.markdown(f'<div class="metric-card"><div class="value">{reviewed}</div>'
                        f'<div class="label">Needs review</div></div>', unsafe_allow_html=True)
            m4.markdown(f'<div class="metric-card"><div class="value">{with_email}</div>'
                        f'<div class="label">Emails found</div></div>', unsafe_allow_html=True)

            st.markdown("###")

            # ── Review expander ──────────────────────────────────────────
            # ── Mail draft count ───────────────────────────────────
            if "Email_Draft" in result.columns:
                drafted = int((result["Email_Draft"].notna() &
                               (result["Email_Draft"] != "")).sum())
                if drafted:
                    st.success(f"✉️ {drafted} personalised email drafts ready — see Email_Draft column below.")

            if reviewed and "Needs_Review" in result.columns:
                flagged = result[result["Needs_Review"] == True]  # noqa
                with st.expander(f"⚠️ {reviewed} rows to double-check"):
                    st.dataframe(flagged, width="stretch")

            # ── Downloads: one place, one button per file ────────────────
            crm_df    = getattr(pipe, "crm_df", None)
            subcat_df = getattr(pipe, "subcat_df", None)
            loc_df    = getattr(pipe, "loc_df", None)
            mail_df   = getattr(pipe, "mail_df", None)

            def _xlsx(df, widths=None):
                """DataFrame -> formatted .xlsx bytes."""
                import openpyxl
                from openpyxl.utils import get_column_letter
                buf = io.BytesIO()
                df.to_excel(buf, index=False, engine="openpyxl")
                buf.seek(0)
                wb = openpyxl.load_workbook(buf)
                ws = wb.active
                widths = widths or {}
                for ci in range(1, ws.max_column + 1):
                    header = ws.cell(1, ci).value
                    ws.column_dimensions[get_column_letter(ci)].width = \
                        widths.get(header, 22)
                    ws.cell(1, ci).font = openpyxl.styles.Font(
                        bold=True, color="FFFFFF")
                    ws.cell(1, ci).fill = openpyxl.styles.PatternFill(
                        "solid", fgColor="0D3B44")
                    for r in range(2, ws.max_row + 1):
                        ws.cell(r, ci).alignment = openpyxl.styles.Alignment(
                            wrap_text=True, vertical="top")
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()

            files = [("📊 Lusha output", result, "leadwhop_output.xlsx", None)]
            if crm_df is not None and len(crm_df):
                files.append(("🗂️ CRM import", crm_df,
                              "crm_import.xlsx", None))
            if subcat_df is not None and len(subcat_df):
                files.append(("🏷️ Sub Category Maps", subcat_df,
                              "sub_category_maps.xlsx", None))
            if loc_df is not None and len(loc_df):
                files.append(("📍 Locations", loc_df,
                              "locations.xlsx", None))
            if mail_df is not None and len(mail_df):
                files.append(("✉️ Mail drafts", mail_df, "mail_drafts.xlsx",
                              {"Name": 22, "Company": 28, "Email": 38,
                               "Email_Subject": 55, "Email_Draft": 90}))

            st.markdown("### Download")
            cols = st.columns(len(files))
            for col, (label, df_out, fname, widths) in zip(cols, files):
                col.download_button(
                    label,
                    _xlsx(df_out, widths),
                    file_name=fname,
                    width="stretch",
                )

            # ── Preview tables ───────────────────────────────────────────
            st.markdown("### Results")
            tabs = st.tabs([label for label, *_ in files])
            for tab, (_, df_out, *_rest) in zip(tabs, files):
                with tab:
                    st.dataframe(df_out, width="stretch", height=350)
